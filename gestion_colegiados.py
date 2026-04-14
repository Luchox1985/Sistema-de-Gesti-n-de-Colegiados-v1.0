from openpyxl import load_workbook
# Abrimos el Excel y apuntamos a la hoja activa
libro = load_workbook(r"C:\Users\Colegiados\Desktop\Informatica\academia_inf\colegiados_prueba.xlsx")
hoja = libro.active
colegiados = []


for fila in hoja.iter_rows(min_row=2, values_only=True):
    colegiados.append({
        "sede":          fila[0],
        "fecha":         fila[1],
        "rut":           fila[2],
        "nombre":        fila[3],
        "genero":        fila[4],
        "email":         fila[5],
        "estado":        fila[6],
        "medio_pago":    fila[7],
        "ultimo_pago":   fila[8],
        "meses_impagos": fila[9]
    })
print(f"Total colegiados cargados: {len(colegiados)}")

opcion = ""

# Muestra el padrón completo — útil para tener el panorama general antes de filtrar
def ver_colegiados(lista):
    print("\n--- LISTA DE COLEGIADOS ---")
    for c in lista:
        print(f"{c['rut']} | {c['nombre']} | {c['estado']}")
    print(f"Total: {len(lista)}")

# Filtra solo los que tienen deuda — los que necesitan gestión de cobranza este mes
def ver_morosos(lista):
    print("\n--- COLEGIADOS MOROSOS ---")
    moroso = 0
    for c in lista:
        if c["estado"] == "MOROSO":
            print(f"{c['rut']} | {c['nombre']} | {c['meses_impagos']} meses impagos" )
            moroso += 1  # cada moroso encontrado suma 1 al contador
    print(f"--------------------------")
    print(f"Total morosos: {moroso}")

# Filtra solo los que están al día — los que van en la nómina del seguro
def ver_activos(lista):
    print("\n--- COLEGIADOS ACTIVOS ---")
    activos = 0
    for c in lista:
        if c["estado"] == "ACTIVO":
            print(f"{c['rut']} | {c['nombre']}")
            activos += 1  # cada activo encontrado suma 1 al contador
    print(f"--------------------------")
    print(f"Total activos: {activos}")

def buscar_rut(lista, rut):
     print("\n--- BÚSQUEDA POR RUT ---")
     for c in lista:
         if c["rut"] == rut:
             print(f"{c['rut']} | {c['nombre']}| {c['sede']} | {c['genero']} | {c['email']} | {c['estado']} | {c['medio_pago']} | {c['meses_impagos']}")

# El sistema corre en loop hasta que el usuario decida salir con 0
while opcion != "0":
    print("==================")
    print("  MENÚ PRINCIPAL  ")
    print("==================")
    print("1. Ver colegiados")
    print("2. Ver morosos")
    print("3. Ver activos")
    print("4. Buscar por RUT")
    print("0. Salir")
    opcion = input("Elige una opción: ")

    # Cada opción deriva a su función — el menú no procesa datos, solo dirige
    if opcion == "1":
        ver_colegiados(colegiados)
    elif opcion == "2":
        ver_morosos(colegiados)
    elif opcion == "3":
        ver_activos(colegiados)
    elif opcion == "4":
        rut_buscar = input("Ingresa el RUT: ")
        buscar_rut(colegiados, rut_buscar)
    elif opcion == "0":
        print("¡Hasta luego!")
    else:
        print("Opción no válida. Intenta de nuevo.")
